#!/usr/bin/env python3
"""Plot Fig. 11 preprocessing overhead as empirical CDFs of SpGEMM runs."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.lines import Line2D
from openpyxl import load_workbook


METHODS = (
    ("runs_6", "RCM"),
    ("runs_7", "Gray"),
    ("runs_1new", "GP"),
    ("runs_2new", "HP"),
    ("runs_4new", "Hierarchical"),
    ("runs_3new", "LSH"),
    ("runs_5", "C-LSH"),
)

METHOD_COLORS = {
    "GP": "#1f77b4",
    "HP": "#ff7f0e",
    "LSH": "#2ca02c",
    "Hierarchical": "#9467bd",
    "C-LSH": "#d62728",
    "RCM": "#b5927d",
    "Gray": "#39454c",
}

METHOD_MARKERS = {
    "GP": "o",
    "HP": "s",
    "LSH": "v",
    "Hierarchical": "^",
    "C-LSH": "d",
    "RCM": "x",
    "Gray": "*",
}

LEGEND_LABELS = {
    "RCM": "RCM",
    "Gray": "Gray",
    "GP": "GP",
    "HP": "HP",
    "Hierarchical": r"Hierarchical $AA^T$",
    "LSH": "Naive LSH",
    "C-LSH": "C-LSH (our work)",
}


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if header is not None and _normalized(header) == target:
            return column
    run_headers = [header for header in headers if "runs" in str(header).lower()]
    raise ValueError(
        f"Cannot find column {name!r}; run columns in worksheet: {run_headers}"
    )


def _to_float_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _read_data(
    workbook_path: Path, sheet_name: str
) -> tuple[dict[str, np.ndarray], int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        name_column = _find_column(headers, "DATA_name")
        run_columns = {
            method: _find_column(headers, column) for column, method in METHODS
        }
        values: dict[str, list[float]] = {method: [] for _, method in METHODS}
        matrix_count = 0

        for row in rows:
            if not any(value is not None for value in row):
                continue
            raw_name = row[name_column]
            if raw_name is None or not str(raw_name).strip():
                continue
            matrix_count += 1
            for method, column in run_columns.items():
                values[method].append(_to_float_or_nan(row[column]))
    finally:
        workbook.close()

    if matrix_count == 0:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")

    return {
        method: np.asarray(method_values, dtype=float)
        for method, method_values in values.items()
    }, matrix_count


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def _plot(values: dict[str, np.ndarray], output_path: Path) -> dict[str, int]:
    _configure_fonts()

    finite_positive = {
        method: method_values[np.isfinite(method_values) & (method_values > 0)]
        for method, method_values in values.items()
    }
    nonempty = [method_values for method_values in finite_positive.values() if len(method_values)]
    if not nonempty:
        raise ValueError("No positive finite preprocessing-run values were found")

    x_max = max(method_values.max() for method_values in nonempty)
    x_min = min(method_values.min() for method_values in nonempty)
    log_base = 5.0
    low_exponent = np.log(x_min) / np.log(log_base)
    high_exponent = np.log(x_max * 1.02) / np.log(log_base)
    x_grid = np.power(
        log_base,
        np.linspace(low_exponent, high_exponent, 800),
    )

    fig, ax = plt.subplots(figsize=(6, 2.8))
    legend_handles = []

    for _, method in METHODS:
        method_values = finite_positive[method]
        if len(method_values) == 0:
            continue

        color = METHOD_COLORS[method]
        marker = METHOD_MARKERS[method]
        proportion = (
            method_values[:, None] <= x_grid[None, :]
        ).sum(axis=0) / len(method_values)
        ax.plot(x_grid, proportion, linewidth=2, color=color, zorder=2)

        sorted_values = np.sort(method_values)
        marker_y = (
            np.searchsorted(sorted_values, sorted_values, side="right")
            / len(sorted_values)
        )
        ax.plot(
            sorted_values,
            marker_y,
            linestyle="None",
            marker=marker,
            markersize=5,
            markeredgecolor=color,
            markerfacecolor="none",
            markeredgewidth=0.6,
            color=color,
            zorder=3,
        )
        legend_handles.append(
            Line2D(
                [0],
                [0],
                color=color,
                linewidth=2,
                marker=marker,
                markersize=5,
                markerfacecolor="none",
                markeredgecolor=color,
                markeredgewidth=0.6,
                label=LEGEND_LABELS[method],
            )
        )

    ax.set_xscale("log", base=5)
    ax.set_xlabel(
        "# of SpGEMM Runs",
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
        labelpad=0.5,
    )
    ax.set_ylabel(
        "Matrix Proportion",
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
        labelpad=0.5,
    )
    ax.set_ylim(0, 1.02)
    ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1.0])
    ax.set_xlim(1.0 / 5.0, x_grid[-1])
    ax.legend(
        handles=legend_handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.98),
        ncol=4,
        fontsize=10,
        frameon=False,
        borderaxespad=0.0,
        handletextpad=0.5,
        labelspacing=0.3,
        handlelength=2,
        columnspacing=1.2,
    )
    ax.grid(True, which="major", linestyle="--", alpha=0.35)
    fig.subplots_adjust(top=0.83)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)

    return {method: len(method_values) for method, method_values in finite_positive.items()}


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot Fig. 11 preprocessing overhead as empirical CDFs."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "Preprocess_Time.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="Xeon5120",
        help="worksheet containing preprocessing data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "Preprocessing_7methods.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    values, matrix_count = _read_data(args.input, args.sheet)
    valid_counts = _plot(values, args.output)
    counts = ", ".join(
        f"{method}={valid_counts[method]}" for _, method in METHODS
    )
    print(f"Plotted {matrix_count} matrices ({counts})")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
