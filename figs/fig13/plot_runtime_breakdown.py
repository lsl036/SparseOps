#!/usr/bin/env python3
"""Plot Fig. 13 normalized runtime breakdown for representative matrices."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook


STAGES = (
    ("setup", "Setup", "#1E5A8F"),
    ("malloc", "Malloc", "#70C870"),
    ("symbolic", "Symbolic", "#FDDA9F"),
    ("dispatcher", "Dispatcher", "#C04030"),
    ("numeric", "Numeric", "#B0D4E3"),
)


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if header is not None and _normalized(header) == target:
            return column
    raise ValueError(f"Cannot find column {name!r}; worksheet headers: {list(headers)}")


def _to_float_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _read_data(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        columns = {
            "DATA_name": _find_column(headers, "DATA_name"),
            "Dense part": _find_column(headers, "Dense part"),
            "sum": _find_column(headers, "sum"),
        }
        columns.update(
            {stage: _find_column(headers, stage) for stage, _, _ in STAGES}
        )

        data = []
        seen_names = set()
        for excel_row, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue
            raw_name = row[columns["DATA_name"]]
            if raw_name is None or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            if name in seen_names:
                raise ValueError(f"Duplicate DATA_name {name!r} in Excel row {excel_row}")
            seen_names.add(name)

            record: dict[str, object] = {
                "DATA_name": name,
                "Dense part": _to_float_or_nan(row[columns["Dense part"]]),
                "sum": _to_float_or_nan(row[columns["sum"]]),
            }
            for stage, _, _ in STAGES:
                record[stage] = _to_float_or_nan(row[columns[stage]])
            data.append(record)
    finally:
        workbook.close()

    if not data:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")

    data.sort(
        key=lambda record: (
            not math.isfinite(float(record["Dense part"])),
            float(record["Dense part"])
            if math.isfinite(float(record["Dense part"]))
            else math.inf,
        )
    )
    return data


def _normalized_stages(
    data: Sequence[dict[str, object]],
) -> tuple[dict[str, np.ndarray], np.ndarray]:
    totals = np.asarray([record["sum"] for record in data], dtype=float)
    valid = np.isfinite(totals) & (totals > 0)
    normalized: dict[str, np.ndarray] = {}

    for stage, _, _ in STAGES:
        stage_values = np.asarray([record[stage] for record in data], dtype=float)
        values = np.zeros(len(data), dtype=float)
        stage_valid = valid & np.isfinite(stage_values)
        values[stage_valid] = stage_values[stage_valid] / totals[stage_valid]
        normalized[stage] = values

    return normalized, valid


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def _plot(
    data: Sequence[dict[str, object]], output_path: Path
) -> tuple[int, float | None, float | None]:
    _configure_fonts()
    normalized, valid = _normalized_stages(data)

    x = np.arange(len(data))
    bottom = np.zeros(len(data), dtype=float)
    fig, ax = plt.subplots(figsize=(6.67, 3))

    for stage, label, color in STAGES:
        values = normalized[stage]
        ax.bar(
            x,
            values,
            bottom=bottom,
            label=label,
            color=color,
            width=0.85,
        )
        bottom += values

    ax.set_xlabel(
        "Matrices (compactness in ascending order)",
        fontdict={"size": 16, "weight": "bold", "family": "serif"},
        labelpad=1,
    )
    ax.set_ylabel(
        "Runtime Fraction",
        fontdict={"size": 16, "weight": "bold", "family": "serif"},
        labelpad=-5,
    )
    ax.set_ylim(0, 1.0)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0, decimals=0))
    ax.legend(
        ncol=5,
        bbox_to_anchor=(0.5, 1.13),
        loc="upper center",
        frameon=False,
        fontsize=9,
    )
    ax.set_xticks(x)
    ax.set_xticklabels([])
    ax.margins(x=0.01)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)

    valid_sums = bottom[valid]
    if len(valid_sums) == 0:
        return 0, None, None
    return int(valid.sum()), float(valid_sums.min()), float(valid_sums.max())


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot Fig. 13 normalized runtime breakdown."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "20_perf_bd.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="Xeon5120",
        help="worksheet containing breakdown data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "20_perf_bd.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    data = _read_data(args.input, args.sheet)
    valid_count, minimum_sum, maximum_sum = _plot(data, args.output)
    print(f"Valid rows: {valid_count}/{len(data)}")
    print(f"Min/max stacked sum (valid): {minimum_sum}/{maximum_sum}")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
