#!/usr/bin/env python3
"""Plot the accumulator compactness experiment directly from hyper_params.xlsx."""

from __future__ import annotations

import argparse
import os
import tempfile
from bisect import bisect_left
from pathlib import Path
from typing import Iterable, Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
from openpyxl import load_workbook


MS_PER_SECOND = 1000.0
DEFAULT_CROSSINGS = (0.16, 0.28, 0.30)


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], candidates: Iterable[str]) -> int:
    normalized_headers = [_normalized(header) for header in headers]
    for candidate in candidates:
        key = _normalized(candidate)
        if key in normalized_headers:
            return normalized_headers.index(key)
    raise ValueError(
        f"Cannot find any of {list(candidates)} in worksheet headers: {list(headers)}"
    )


def _find_density_column(headers: Sequence[object], value_column: int) -> int:
    for column in range(value_column - 1, -1, -1):
        if _normalized(headers[column]) == "density":
            return column
    raise ValueError(f"Cannot find a Density column before column {value_column + 1}")


def _read_platform_data(
    workbook_path: Path, sheet_name: str
) -> list[tuple[list[float], list[float], list[float]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        column_specs = (
            (("Hash Time (s)",), ("Dense Time (s)",)),
            (("Hash Time_2", "Hash Time 2"), ("Dense Time_2", "Dense Time 2")),
            (("Hash Time_3", "Hash Time 3"), ("Dense Time_3", "Dense Time 3")),
        )
        indices = []
        for hash_candidates, dense_candidates in column_specs:
            hash_column = _find_column(headers, hash_candidates)
            dense_column = _find_column(headers, dense_candidates)
            density_column = _find_density_column(headers, hash_column)
            indices.append((density_column, hash_column, dense_column))

        platforms = [([], [], []) for _ in indices]
        for excel_row, row in enumerate(rows, start=2):
            for platform, (density_column, hash_column, dense_column) in zip(
                platforms, indices
            ):
                values = (row[density_column], row[hash_column], row[dense_column])
                if all(value is None for value in values):
                    continue
                if any(value is None for value in values):
                    raise ValueError(
                        f"Incomplete compactness data in Excel row {excel_row}: {values}"
                    )
                platform[0].append(float(values[0]))
                platform[1].append(float(values[1]) * MS_PER_SECOND)
                platform[2].append(float(values[2]) * MS_PER_SECOND)
    finally:
        workbook.close()

    reference_density = platforms[0][0]
    if not reference_density:
        raise ValueError(f"Worksheet {sheet_name!r} contains no compactness data")
    if any(right <= left for left, right in zip(reference_density, reference_density[1:])):
        raise ValueError("Density values must be strictly increasing")

    for platform_number, (density, _, _) in enumerate(platforms[1:], start=2):
        if len(density) != len(reference_density) or any(
            abs(left - right) > 1.0e-12
            for left, right in zip(reference_density, density)
        ):
            raise ValueError(
                f"Platform {platform_number} Density values do not match platform 1"
            )

    return platforms


def _interpolate(x_query: float, x: Sequence[float], y: Sequence[float]) -> float:
    if x_query < x[0] or x_query > x[-1]:
        raise ValueError(
            f"Crossing marker {x_query} is outside Density range [{x[0]}, {x[-1]}]"
        )

    right = bisect_left(x, x_query)
    if right == 0 or x[right] == x_query:
        return y[right]

    left = right - 1
    weight = (x_query - x[left]) / (x[right] - x[left])
    return y[left] + weight * (y[right] - y[left])


def _plot(
    platforms: Sequence[tuple[list[float], list[float], list[float]]],
    crossings: Sequence[float],
    output_path: Path,
) -> list[tuple[float, float]]:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )

    styles = (
        ("#2E5A88", "#66B7D0", "-", 1.0),
        ("#D4A017", "#D46D3A", "-.", 0.8),
        ("#6CC788", "#3E6A3D", ":", 0.85),
    )

    fig, ax = plt.subplots(figsize=(5, 2))
    hash_lines = []
    dense_lines = []
    marker_points = []

    for platform_number, ((density, hash_ms, dense_ms), style, crossing) in enumerate(
        zip(platforms, styles, crossings), start=1
    ):
        hash_color, dense_color, line_style, alpha = style
        (hash_line,) = ax.plot(
            density,
            hash_ms,
            color=hash_color,
            alpha=alpha,
            marker="o",
            linewidth=1.5,
            markersize=3,
            linestyle=line_style,
            label="Hash",
        )
        (dense_line,) = ax.plot(
            density,
            dense_ms,
            color=dense_color,
            alpha=alpha,
            marker="D",
            linewidth=1.5,
            markersize=3,
            linestyle=line_style,
            label=f"Block-SPA (plat. {platform_number})",
        )
        hash_lines.append(hash_line)
        dense_lines.append(dense_line)

        hash_at_crossing = _interpolate(crossing, density, hash_ms)
        dense_at_crossing = _interpolate(crossing, density, dense_ms)
        marker_points.append((crossing, (hash_at_crossing + dense_at_crossing) / 2.0))

    marker_x, marker_y = zip(*marker_points)
    ax.scatter(
        marker_x,
        marker_y,
        marker="x",
        s=50,
        c="darkred",
        linewidths=1.2,
        zorder=6,
    )

    label_font = {"size": 14, "weight": "bold", "family": "serif"}
    ax.set_xlabel("Compactness", fontdict=label_font, labelpad=0.5)
    ax.set_ylabel(
        "Time (ms)",
        fontdict={"size": 12, "weight": "bold", "family": "serif"},
        labelpad=-1,
    )
    plt.setp(ax.get_xticklabels(), fontsize=11, fontweight="bold")
    plt.setp(ax.get_yticklabels(), fontsize=11, fontweight="bold")

    handles = tuple(hash_lines + dense_lines)
    ax.legend(
        handles,
        [line.get_label() for line in handles],
        ncol=2,
        loc="best",
        markerscale=0.8,
        fontsize=8,
        columnspacing=0.8,
    )
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)
    return marker_points


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot the three-platform compactness experiment as a PDF."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "hyper_params.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="compactness",
        help="worksheet containing compactness data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "Fig6_Compactness.pdf",
        help="output PDF path (default: %(default)s)",
    )
    parser.add_argument(
        "--crossings",
        type=float,
        nargs=3,
        metavar=("PLAT1", "PLAT2", "PLAT3"),
        default=DEFAULT_CROSSINGS,
        help="x coordinates of the three crossing markers (default: 0.16 0.28 0.30)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    platforms = _read_platform_data(args.input, args.sheet)
    marker_points = _plot(platforms, args.crossings, args.output)

    print(f"Saved {args.output.resolve()}")
    for platform_number, (x_value, y_value) in enumerate(marker_points, start=1):
        print(
            f"Platform {platform_number} marker: compactness={x_value:.4f}, "
            f"time={y_value:.3f} ms"
        )


if __name__ == "__main__":
    main()
