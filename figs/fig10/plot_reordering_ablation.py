#!/usr/bin/env python3
"""Plot the Fig. 10 reordering and accumulator ablation results."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Patch
from openpyxl import load_workbook


PERF_COLUMNS = (
    "GP_Hash",
    "GP_HYB",
    "HP_Hash",
    "HP_HYB",
    "RCM_Hash",
    "RCM_HYB",
    "Gray_Hash",
    "Gray_HYB",
    "AAT_Hash",
    "AAT_HYB",
    "LSH_Hash",
    "LSH_HYB",
    "CLSH_Hash",
    "CLSH_HYB",
)

METHODS = (
    ("RCM_Hash", "RCM_HYB", "Reverse Cuthill-McKee (RCM)"),
    ("Gray_Hash", "Gray_HYB", "Gray Code Ordering (Gray)"),
    ("GP_Hash", "GP_HYB", "Graph Partitioning (GP)"),
    ("HP_Hash", "HP_HYB", "Hypergraph Partitioning (HP)"),
    ("AAT_Hash", "AAT_HYB", r"Hierarchical $AA^T$ Clustering"),
    ("LSH_Hash", "LSH_HYB", "Naive LSH"),
    ("CLSH_Hash", "CLSH_HYB", "C-LSH (our work)"),
)

COLORS = (
    "#b5927d",
    "#39454c",
    "#80a8e2",
    "#F2C867",
    "#9b5894",
    "#2ca02c",
    "#FF4D4D",
)


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if header is not None and _normalized(header) == target:
            return column
    raise ValueError(f"Cannot find column {name!r}; worksheet headers: {list(headers)}")


def _to_float_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _read_data(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        columns = {
            "mtx_ID": _find_column(headers, "mtx_ID"),
            "DATA_name": _find_column(headers, "DATA_name"),
            "Dense part": _find_column(headers, "Dense part"),
        }
        columns.update({name: _find_column(headers, name) for name in PERF_COLUMNS})

        data = []
        seen_names = set()
        for excel_row, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue

            raw_name = row[columns["DATA_name"]]
            if raw_name is None or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            if name in seen_names:
                raise ValueError(f"Duplicate DATA_name {name!r} in Excel row {excel_row}")
            seen_names.add(name)

            record: dict[str, object] = {
                "mtx_ID": _to_float_or_nan(row[columns["mtx_ID"]]),
                "DATA_name": name,
                "Dense part": _to_float_or_nan(row[columns["Dense part"]]),
            }
            for column in PERF_COLUMNS:
                record[column] = _to_float_or_nan(row[columns[column]])
            data.append(record)
    finally:
        workbook.close()

    if not data:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")

    # Match the notebook's stable mtx_ID sort followed by stable Dense part sort.
    data.sort(
        key=lambda record: (
            not math.isfinite(float(record["mtx_ID"])),
            float(record["mtx_ID"])
            if math.isfinite(float(record["mtx_ID"]))
            else math.inf,
        )
    )
    data.sort(
        key=lambda record: (
            not math.isfinite(float(record["Dense part"])),
            float(record["Dense part"])
            if math.isfinite(float(record["Dense part"]))
            else math.inf,
        )
    )
    return data


def _ellipsize(value: object, max_length: int = 8) -> str:
    text = str(value)
    if len(text) <= max_length:
        return text
    return text[: max_length - 3] + "..."


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "hatch.linewidth": 1.0,
        }
    )


def _plot(data: Sequence[dict[str, object]], output_path: Path) -> int:
    _configure_fonts()

    x = np.arange(len(data))
    names = [str(record["DATA_name"]) for record in data]
    group_width = 0.88
    bar_width = group_width / len(METHODS)
    offsets = (
        np.arange(len(METHODS)) - (len(METHODS) - 1) / 2.0
    ) * bar_width
    bar_fill_ratio = 0.86
    zero_x_positions: list[float] = []

    fig, ax = plt.subplots(figsize=(20, 3.5))

    for method_index, ((hash_col, hybrid_col, _), offset) in enumerate(
        zip(METHODS, offsets)
    ):
        hash_values = np.asarray([row[hash_col] for row in data], dtype=float)
        hybrid_values = np.asarray([row[hybrid_col] for row in data], dtype=float)
        positions = x + offset
        width = bar_width * bar_fill_ratio

        invalid = (
            ~np.isfinite(hash_values)
            | ~np.isfinite(hybrid_values)
            | np.isclose(hash_values, 0.0)
            | np.isclose(hybrid_values, 0.0)
        )
        zero_x_positions.extend(positions[invalid].tolist())

        line_width = 1.2
        valid = ~invalid
        ax.bar(
            positions[valid],
            hash_values[valid],
            width=width,
            align="center",
            color=COLORS[method_index],
            edgecolor=COLORS[method_index],
            linewidth=line_width,
            zorder=2,
        )

        difference = hybrid_values - hash_values
        valid_delta = ~invalid & np.isfinite(difference) & ~np.isclose(difference, 0.0)
        positive = valid_delta & (difference > 0)
        negative = valid_delta & (difference < 0)

        positive_hatch = {
            "width": width,
            "align": "center",
            "hatch": "xx",
            "facecolor": "white",
            "edgecolor": COLORS[method_index],
            "linewidth": line_width,
            "zorder": 3,
        }
        negative_hatch = {
            "width": width,
            "align": "center",
            "hatch": "xx",
            "facecolor": COLORS[method_index],
            "edgecolor": "black",
            "linewidth": line_width,
            "zorder": 3,
        }
        if np.any(positive):
            ax.bar(
                positions[positive],
                difference[positive],
                bottom=hash_values[positive],
                **positive_hatch,
            )
        if np.any(negative):
            ax.bar(
                positions[negative],
                difference[negative],
                bottom=hash_values[negative],
                **negative_hatch,
            )

    for patch in ax.patches:
        patch.set_snap(False)

    ax.relim()
    ax.autoscale_view()

    for position in zero_x_positions:
        ax.text(
            float(position),
            -5,
            "X",
            fontsize=12,
            color="darkred",
            ha="center",
            va="bottom",
            zorder=5,
            clip_on=False,
        )

    ax.set_xticks(x)
    ax.set_xticklabels([_ellipsize(name) for name in names], fontsize=12)
    ax.set_xlabel(
        "Representative Matrices (the proportion of compact clusters in ascending order)",
        fontdict={"size": 18, "weight": "bold", "family": "serif"},
    )
    ax.set_ylabel(
        "Performance (GFlops)",
        fontdict={"size": 18, "weight": "bold", "family": "serif"},
    )
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.set_axisbelow(True)

    x_padding = 1.5 * bar_width
    drawn_width = bar_width * bar_fill_ratio
    ax.set_xlim(
        x[0] + offsets.min() - drawn_width / 2.0 - x_padding,
        x[-1] + offsets.max() + drawn_width / 2.0 + x_padding,
    )

    for boundary in range(len(x) - 1):
        ax.axvline(
            boundary + 0.5,
            color="0.55",
            linestyle=":",
            linewidth=0.9,
            alpha=0.75,
            zorder=1,
        )

    legend_line_width = 1.2
    method_handles = [
        Patch(
            facecolor=COLORS[index],
            edgecolor=COLORS[index],
            linewidth=legend_line_width,
            label=method_name,
        )
        for index, (_, _, method_name) in enumerate(METHODS)
    ]
    fig.legend(
        handles=method_handles,
        loc="lower center",
        bbox_to_anchor=(0.51, 0.93),
        ncol=7,
        fontsize=12,
        frameon=False,
        columnspacing=1.2,
        handletextpad=0.8,
        handlelength=2,
    )

    inner_handles = [
        Patch(
            facecolor="gray",
            edgecolor="gray",
            linewidth=legend_line_width,
            label="Hash Accumulator",
        ),
        Patch(
            facecolor="white",
            edgecolor="gray",
            linewidth=legend_line_width,
            hatch="xxx",
            label="Hybrid Accumulation",
        ),
    ]
    ax.legend(
        handles=inner_handles,
        loc="upper left",
        fontsize=14,
        framealpha=0.92,
        columnspacing=1.0,
        handletextpad=0.8,
        handlelength=2,
        borderpad=0.5,
        labelspacing=0.5,
    )

    fig.subplots_adjust(top=0.88)
    fig.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)
    return len(zero_x_positions)


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot the Fig. 10 reordering and accumulator ablation results."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "ReorderDraw.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="Sheet1",
        help="worksheet containing ablation data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "Ablation0709.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    data = _read_data(args.input, args.sheet)
    invalid_bars = _plot(data, args.output)
    print(f"Plotted {len(data)} matrices with {invalid_bars} invalid method bars")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
