#!/usr/bin/env python3
"""Plot Fig. 12 preprocessing memory overhead for seven reorderings."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib.lines import Line2D
from openpyxl import load_workbook


METHODS = (
    ("RCM_overhead(GB)", "RCM"),
    ("Gray_overhead(GB)", "Gray"),
    ("GP_overhead(GB)", "GP"),
    ("HP_overhead(GB)", "HP"),
    ("Hier Memory (GB)", "Hierarchical"),
    ("LSH Overhead(GB)", "LSH"),
    ("C-LSH_Mem(GB)", "C-LSH"),
)

METHOD_COLORS = {
    "RCM": "#b5927d",
    "Gray": "#39454c",
    "GP": "#1f77b4",
    "HP": "#ff7f0e",
    "Hierarchical": "#9467bd",
    "LSH": "#2ca02c",
    "C-LSH": "#d62728",
}

METHOD_MARKERS = {
    "RCM": "x",
    "Gray": "*",
    "GP": "o",
    "HP": "s",
    "Hierarchical": "^",
    "LSH": "v",
    "C-LSH": "d",
}

LEGEND_LABELS = {
    "RCM": "RCM",
    "Gray": "Gray",
    "GP": "GP",
    "HP": "HP",
    "Hierarchical": r"Hierarchical $AA^T$",
    "LSH": "Naive LSH",
    "C-LSH": "C-LSH (our work)",
}

MIB_IN_GIB = 1.0 / 1024.0
MEMORY_GIB = np.asarray([MIB_IN_GIB, 0.5, 1.0, 512.0, 1024.0])
X_POSITIONS = np.asarray([0.0, 1.0, 2.0, 3.0, 4.0])
X_TICK_LABELS = ("1MB", "0.5GB", "1GB", "0.5TB", "1TB")
LOG_MEMORY_GIB = np.log(MEMORY_GIB)
X_LEFT = X_POSITIONS[0] - 0.08
X_RIGHT = X_POSITIONS[-1] + 0.08


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if header is not None and _normalized(header) == target:
            return column
    raise ValueError(f"Cannot find column {name!r}; worksheet headers: {list(headers)}")


def _to_float_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _read_data(
    workbook_path: Path, sheet_name: str
) -> tuple[dict[str, np.ndarray], int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        name_column = _find_column(headers, "DATA_name")
        memory_columns = {
            method: _find_column(headers, column) for column, method in METHODS
        }
        values: dict[str, list[float]] = {method: [] for _, method in METHODS}
        matrix_count = 0

        for row in rows:
            if not any(value is not None for value in row):
                continue
            raw_name = row[name_column]
            if raw_name is None or not str(raw_name).strip():
                continue
            matrix_count += 1
            for method, column in memory_columns.items():
                values[method].append(_to_float_or_nan(row[column]))
    finally:
        workbook.close()

    if matrix_count == 0:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")

    return {
        method: np.asarray(method_values, dtype=float)
        for method, method_values in values.items()
    }, matrix_count


def _gib_to_equal_space(values_gib: np.ndarray) -> np.ndarray:
    values_gib = np.asarray(values_gib, dtype=float)
    values_gib = np.maximum(values_gib, np.finfo(float).tiny)
    return np.interp(np.log(values_gib), LOG_MEMORY_GIB, X_POSITIONS)


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def _plot(values: dict[str, np.ndarray], output_path: Path) -> dict[str, int]:
    _configure_fonts()

    finite_positive = {
        method: method_values[np.isfinite(method_values) & (method_values > 0)]
        for method, method_values in values.items()
    }
    if not any(len(method_values) for method_values in finite_positive.values()):
        raise ValueError("No positive finite preprocessing-memory values were found")

    fig, ax = plt.subplots(figsize=(6, 2.5))
    legend_handles = []

    for _, method in METHODS:
        method_values = finite_positive[method]
        count = len(method_values)
        if count == 0:
            continue

        sorted_gib = np.sort(method_values)
        x_values = _gib_to_equal_space(sorted_gib)
        y_values = np.arange(1, count + 1, dtype=float) / count
        plot_x = np.concatenate(
            ([X_LEFT, x_values[0], x_values[0]], x_values[1:], [X_RIGHT])
        )
        plot_y = np.concatenate(
            ([0.0, 0.0, y_values[0]], y_values[1:], [y_values[-1]])
        )

        color = METHOD_COLORS[method]
        marker = METHOD_MARKERS[method]
        marker_indices = list(range(2, count + 2))
        ax.plot(
            plot_x,
            plot_y,
            linewidth=2,
            color=color,
            marker=marker,
            markersize=5,
            markevery=marker_indices,
            markeredgecolor=color,
            markerfacecolor="none",
            markeredgewidth=0.6,
        )
        legend_handles.append(
            Line2D(
                [0],
                [0],
                color=color,
                linewidth=2,
                marker=marker,
                markersize=5,
                markerfacecolor="none",
                markeredgecolor=color,
                markeredgewidth=0.6,
                label=LEGEND_LABELS[method],
            )
        )

    ax.set_xticks(X_POSITIONS)
    ax.set_xticklabels(X_TICK_LABELS)
    ax.set_xlabel(
        "Required Memory Space",
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
        labelpad=0.5,
    )
    ax.set_ylabel(
        "Matrix Proportion",
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
        labelpad=0.5,
    )
    ax.set_ylim(0, 1.02)
    ax.set_xlim(X_LEFT, X_RIGHT)
    ax.legend(
        handles=legend_handles,
        loc="lower right",
        ncol=1,
        fontsize=10,
        frameon=True,
        framealpha=0.95,
        borderaxespad=0.5,
        handletextpad=0.5,
        labelspacing=0.3,
        handlelength=2,
        columnspacing=1.5,
    )
    ax.grid(True, which="major", linestyle="--", alpha=0.35)
    fig.subplots_adjust(top=0.83)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)

    return {method: len(method_values) for method, method_values in finite_positive.items()}


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot Fig. 12 preprocessing memory overhead for seven methods."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "Preprocess_Mem.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="AMDEPYC",
        help="worksheet containing memory data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "Preprocessing_Mem_7methods.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    values, matrix_count = _read_data(args.input, args.sheet)
    valid_counts = _plot(values, args.output)
    counts = ", ".join(
        f"{method}={valid_counts[method]}" for _, method in METHODS
    )
    print(f"Plotted {matrix_count} matrices ({counts})")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
