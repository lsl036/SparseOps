#!/usr/bin/env python3
"""Plot L2 hit rate and normalized eviction volume from hyper_params.xlsx."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
from matplotlib.ticker import FixedLocator, PercentFormatter
from openpyxl import load_workbook


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if _normalized(header) == target:
            return column
    raise ValueError(f"Cannot find column {name!r}; worksheet headers: {list(headers)}")


def _read_l2_data(
    workbook_path: Path, sheet_name: str
) -> tuple[list[float], list[float], list[float]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        param_column = 0
        hit_column = _find_column(headers, "hit_ave")
        evict_column = _find_column(headers, "evict_ratio")
        params: list[float] = []
        hit_rates: list[float] = []
        evict_ratios: list[float] = []

        for excel_row, row in enumerate(rows, start=2):
            values = (row[param_column], row[hit_column], row[evict_column])
            if all(value is None for value in values):
                continue
            if any(value is None for value in values):
                raise ValueError(f"Incomplete L2 data in Excel row {excel_row}: {values}")
            try:
                parsed = tuple(float(value) for value in values)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"Non-numeric L2 data in Excel row {excel_row}: {values}"
                ) from exc
            if not all(math.isfinite(value) for value in parsed):
                raise ValueError(f"Non-finite L2 data in Excel row {excel_row}: {values}")
            params.append(parsed[0])
            hit_rates.append(parsed[1])
            evict_ratios.append(parsed[2])
    finally:
        workbook.close()

    if not params:
        raise ValueError(f"Worksheet {sheet_name!r} contains no L2 data")
    if any(right <= left for left, right in zip(params, params[1:])):
        raise ValueError("L2 cache budget values must be strictly increasing")
    if any(value < 0.0 or value > 1.0 for value in hit_rates):
        raise ValueError("L2 hit rates must be in the range [0, 1]")

    return params, hit_rates, evict_ratios


def _plot(
    params: Sequence[float],
    hit_rates: Sequence[float],
    evict_ratios: Sequence[float],
    output_path: Path,
) -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )

    color_hit = "#2E5A88"
    color_evict = "#D46D3A"
    fig, ax = plt.subplots(figsize=(6, 3))

    (line_hit,) = ax.plot(
        params,
        hit_rates,
        color=color_hit,
        marker="o",
        label="L2 Cache Hit Rate",
        linewidth=3,
        markersize=7,
    )
    ax.set_xlabel(
        "L2 Cache Budget",
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
    )
    ax.xaxis.set_major_locator(FixedLocator(list(params)))
    ax.set_ylabel(
        "L2 Cache Hit Rate",
        color=color_hit,
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
    )
    ax.tick_params(axis="y", labelcolor=color_hit)
    ax.set_ylim(0.5, 0.7)
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))
    ax.yaxis.set_major_locator(FixedLocator([0.5, 0.55, 0.6, 0.65, 0.7]))

    ax2 = ax.twinx()
    (line_evict,) = ax2.plot(
        params,
        evict_ratios,
        color=color_evict,
        marker="s",
        label="L2-to-L3 Eviction Volume (Normalized)",
        linewidth=3,
        markersize=7,
    )
    ax2.set_ylabel(
        "Normalized Eviction Volume",
        color=color_evict,
        fontdict={"size": 14, "weight": "bold", "family": "serif"},
    )
    ax2.tick_params(axis="y", labelcolor=color_evict)
    ax2.set_ylim(1.0, 1.4)
    ax2.yaxis.set_major_locator(FixedLocator([1.0, 1.1, 1.2, 1.3, 1.4]))

    tick_style = {"fontsize": 12, "fontweight": "bold"}
    plt.setp(ax.get_xticklabels(), **tick_style)
    plt.setp(ax.get_yticklabels(), **tick_style)
    plt.setp(ax2.get_yticklabels(), **tick_style)

    for budget in (0.6, 0.8):
        ax.axvline(
            budget,
            color="black",
            linestyle=":",
            linewidth=0.8,
            zorder=0,
        )
    ax.axvline(
        0.7,
        color="black",
        linestyle="--",
        linewidth=1.0,
        zorder=0,
    )

    ax.legend(
        [line_hit, line_evict],
        [line_hit.get_label(), line_evict.get_label()],
        loc="upper center",
        fontsize=9,
    )
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot L2 hit rate and normalized eviction volume as a PDF."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir.parent / "fig6" / "hyper_params.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="L2_hitrate",
        help="worksheet containing L2 data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "L2_budget.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    params, hit_rates, evict_ratios = _read_l2_data(args.input, args.sheet)
    _plot(params, hit_rates, evict_ratios, args.output)

    print(f"Read {len(params)} L2 budget points from {args.input.resolve()}")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
