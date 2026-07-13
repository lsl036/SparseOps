#!/usr/bin/env python3
"""Plot three-platform SpGEMM performance and LeSpGEMM speedup."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Mapping, Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from matplotlib import ticker
from openpyxl import load_workbook


METHODS = ("Ha", "MKL", "kokkos", "CSeg", "Hier", "Gra", "Le")
METHOD_TITLES = {
    "Ha": "HaSpGEMM",
    "MKL": "MKL / AOCL Sparse",
    "kokkos": "Kokkos-Kernels",
    "CSeg": "CSeg",
    "Hier": "Hierarchical SpGEMM",
    "Gra": "SuiteSparse:GraphBLAS",
    "Le": "LeSpGEMM (our work)",
}
COLORS = ("#1F70A9", "#E36D4F", "#7AA35A")
PLATFORM_LABELS = (
    "Intel Xeon 5120",
    "Intel Xeon 6248",
    "AMD EPYC 7C13",
)

# Coefficients are preserved from the paper notebook: y = slope * log10(FLOPs) + intercept.
FIT_COEFFICIENTS = {
    "Ha": (
        (0.38431091, -2.05135662),
        (1.29089978, -8.16894910),
        (0.56324477, -2.99689812),
    ),
    "MKL": (
        (1.88399042, -9.93865368),
        (0.31097100, -0.44944435),
        (3.55717504, -15.11710157),
    ),
    "kokkos": (
        (1.81591279, -10.53649912),
        (2.69282277, -15.02884474),
        (3.81754960, -23.95329542),
    ),
    "CSeg": (
        (3.14114274, -17.04523299),
        (1.69476936, -7.79907939),
        (8.02550542, -51.18721824),
    ),
    "Hier": (
        (1.97303820, -10.55178833),
        (2.75598810, -15.86923140),
        (7.84864908, -44.64196811),
    ),
    "Gra": (
        (5.50000000, -38.35741187),
        (7.50000000, -53.37277420),
        (13.92855477, -95.66043826),
    ),
    "Le": (
        (6.50000000, -42.35741187),
        (8.75000000, -55.69752675),
        (14.67807529, -88.14799672),
    ),
}


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], candidates: Sequence[str]) -> int:
    normalized_headers = [_normalized(header) for header in headers]
    for candidate in candidates:
        key = _normalized(candidate)
        if key in normalized_headers:
            return normalized_headers.index(key)
    raise ValueError(
        f"Cannot find any of {list(candidates)} in worksheet headers: {list(headers)}"
    )


def _read_sheet(
    workbook_path: Path,
    sheet_name: str,
    method_columns: Mapping[str, Sequence[str]],
) -> tuple[list[str], dict[str, dict[str, float]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist in {workbook_path}; "
                f"available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        columns = {
            "DATA_name": _find_column(headers, ("DATA_name",)),
            "FLOPs": _find_column(headers, ("FLOPs",)),
        }
        for method, candidates in method_columns.items():
            columns[method] = _find_column(headers, candidates)

        names: list[str] = []
        records: dict[str, dict[str, float]] = {}
        for excel_row, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue
            raw_name = row[columns["DATA_name"]]
            if raw_name is None or not str(raw_name).strip():
                raise ValueError(f"Missing DATA_name in {sheet_name} row {excel_row}")
            name = str(raw_name).strip()
            if name in records:
                raise ValueError(f"Duplicate DATA_name {name!r} in {sheet_name}")

            record: dict[str, float] = {}
            for key, column in columns.items():
                if key == "DATA_name":
                    continue
                value = row[column]
                try:
                    number = float(value)
                except (TypeError, ValueError) as exc:
                    raise ValueError(
                        f"Non-numeric {key} value in {sheet_name} row {excel_row}: {value!r}"
                    ) from exc
                if not math.isfinite(number):
                    raise ValueError(
                        f"Non-finite {key} value in {sheet_name} row {excel_row}: {value!r}"
                    )
                record[key] = number
            names.append(name)
            records[name] = record
    finally:
        workbook.close()

    if not names:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")
    return names, records


def _aligned_values(
    records: Mapping[str, Mapping[str, float]], names: Sequence[str], column: str
) -> np.ndarray:
    missing = [name for name in names if name not in records]
    extra = [name for name in records if name not in set(names)]
    if missing or extra:
        raise ValueError(
            f"DATA_name mismatch: missing={missing[:5]}, extra={extra[:5]}"
        )
    return np.asarray([records[name][column] for name in names], dtype=float)


def _pow10_exponent(value: float, _: int) -> str:
    exponent = int(round(value))
    if abs(value - exponent) > 1.0e-6:
        return ""
    return rf"$10^{{{exponent}}}$"


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def _plot(
    names: Sequence[str],
    xeon5120: Mapping[str, Mapping[str, float]],
    xeon6248: Mapping[str, Mapping[str, float]],
    amd: Mapping[str, Mapping[str, float]],
    output_path: Path,
) -> None:
    _configure_fonts()

    flops = _aligned_values(xeon5120, names, "FLOPs")
    if np.any(flops <= 0):
        raise ValueError("FLOPs values must be positive before applying log10")
    order = np.argsort(flops)
    log_flops = np.log10(flops[order])

    platform_records = (xeon5120, xeon6248, amd)
    method_values = {
        method: tuple(
            _aligned_values(records, names, method) for records in platform_records
        )
        for method in METHODS
    }
    le_values = method_values["Le"]

    fig, axes = plt.subplots(
        2,
        7,
        figsize=(18, 5),
        sharex=True,
        gridspec_kw={"height_ratios": [3, 2]},
    )

    for axis in axes.flat:
        axis.xaxis.set_major_locator(ticker.MaxNLocator(nbins=5, integer=True))
        axis.xaxis.set_major_formatter(ticker.FuncFormatter(_pow10_exponent))
        axis.grid(True, linestyle="-", alpha=0.35)
        axis.tick_params(labelsize=10)
        plt.setp(axis.get_xticklabels(), fontweight="bold")
        plt.setp(axis.get_yticklabels(), fontweight="bold")

    scatter_handles = []
    for column, method in enumerate(METHODS):
        axis = axes[0, column]
        values = method_values[method]
        scatters = (
            axis.scatter(
                log_flops,
                values[0][order],
                s=8,
                label=PLATFORM_LABELS[0],
                marker="s",
                color=COLORS[0],
                linewidth=0,
                alpha=0.6,
            ),
            axis.scatter(
                log_flops,
                values[1][order],
                s=13,
                label=PLATFORM_LABELS[1],
                marker="^",
                color=COLORS[1],
                linewidth=0,
                alpha=0.6,
            ),
            axis.scatter(
                log_flops,
                values[2][order],
                s=13,
                label=PLATFORM_LABELS[2],
                marker="+",
                color=COLORS[2],
                linewidth=1,
                alpha=0.6,
            ),
        )
        if column == 0:
            scatter_handles = list(scatters)

        for platform, (slope, intercept) in enumerate(FIT_COEFFICIENTS[method]):
            axis.plot(
                log_flops,
                slope * log_flops + intercept,
                linestyle="-",
                linewidth=1.4,
                color=COLORS[platform],
                label="_nolegend_",
            )

        axis.set_title(METHOD_TITLES[method], fontsize=12, fontweight="bold", pad=5)
        axis.set_ylim(0, 80)
        axis.set_yticks([20, 40, 60, 80])
        if column == 0:
            axis.set_ylabel(
                "Performance (GFlops)",
                fontdict={"size": 16, "weight": "bold", "family": "serif"},
            )
            axis.legend(
                handles=scatter_handles,
                loc="upper left",
                fontsize=10,
                markerscale=2,
                borderaxespad=0.2,
            )
        else:
            axis.tick_params(axis="y", labelleft=False)

    for column, method in enumerate(METHODS):
        axis = axes[1, column]
        values = method_values[method]
        if method == "Le":
            speedups = tuple(np.ones_like(log_flops) for _ in platform_records)
            speedup_x = (log_flops, log_flops, log_flops)
        else:
            speedups_list = []
            speedup_x_list = []
            for baseline, le_value in zip(values, le_values):
                valid = baseline > 0
                speedups_list.append(le_value[valid] / baseline[valid])
                speedup_x_list.append(np.log10(flops[valid]))
            speedups = tuple(speedups_list)
            speedup_x = tuple(speedup_x_list)

        axis.scatter(
            speedup_x[0],
            speedups[0],
            s=8,
            marker="s",
            color=COLORS[0],
            linewidth=0,
            alpha=0.6,
        )
        axis.scatter(
            speedup_x[1],
            speedups[1],
            s=13,
            marker="^",
            color=COLORS[1],
            linewidth=0,
            alpha=0.6,
        )
        axis.scatter(
            speedup_x[2],
            speedups[2],
            s=13,
            marker="+",
            color=COLORS[2],
            linewidth=1,
            alpha=0.6,
        )
        axis.axhline(1, linestyle="--", linewidth=1.5, color="#B22222")
        axis.set_ylim(0, 10)
        axis.set_yticks([0, 1, 2, 4, 6, 8, 10])
        if column == 0:
            axis.set_ylabel(
                "Speedup",
                fontdict={"size": 16, "weight": "bold", "family": "serif"},
            )
        else:
            axis.tick_params(axis="y", labelleft=False)

    fig.supxlabel(
        "Number of Floating-Point Operations",
        fontsize=18,
        fontweight="bold",
        family="serif",
        y=0.02,
    )
    fig.tight_layout(rect=[0, 0.07, 1, 1])
    fig.subplots_adjust(hspace=0.08, wspace=0.08)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot three-platform SpGEMM performance and speedup."
    )
    parser.add_argument(
        "--benchmarks",
        type=Path,
        default=script_dir / "benchmarks.xlsx",
        help="Intel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--amd",
        type=Path,
        default=script_dir / "AMD_draw.xlsx",
        help="AMD workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "Performance600_3platform.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    intel_columns = {
        "Ha": ("Ha_GFlops",),
        "MKL": ("MKL_GFlops",),
        "kokkos": ("kokkos_GFlops",),
        "CSeg": ("CSeg_GFlops",),
        "Hier": ("Hier_GFlops",),
        "Gra": ("Gra_GFlops",),
        "Le": ("Le_GFlops",),
    }
    amd_columns = {
        **intel_columns,
        "MKL": ("AOCL_GFlops", "AOCL_Gflops"),
    }

    names, xeon5120 = _read_sheet(args.benchmarks, "Xeon5120", intel_columns)
    _, xeon6248 = _read_sheet(args.benchmarks, "Xeon6248", intel_columns)
    _, amd = _read_sheet(args.amd, "Sheet1", amd_columns)
    _plot(names, xeon5120, xeon6248, amd, args.output)

    print(f"Aligned and plotted {len(names)} datasets")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
