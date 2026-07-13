#!/usr/bin/env python3
"""Plot grouped SpGEMM performance bars for the 20 representative matrices."""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Sequence

os.environ.setdefault(
    "MPLCONFIGDIR",
    str(Path(tempfile.gettempdir()) / f"sparseops-matplotlib-{os.getuid()}"),
)

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


PERF_COLUMNS = (
    "Ha_GFlops",
    "MKL_GFlops",
    "kokkos_GFlops",
    "CSeg_Gflops",
    "Hier_GFlops",
    "Gra_GFlops",
    "Le_GFlops",
)
COLORS = (
    "#D9C6B3",
    "#2E5A88",
    "#8DB67A",
    "#F2C867",
    "#B5A0C2",
    "#7EC0EE",
    "#FF4D4D",
)
LABELS = (
    "HaSpGEMM",
    "MKL SpBLAS",
    "Kokkos-Kernel",
    "CSeg",
    "Hierarchical SpGEMM",
    "SuiteSparse:GraphBLAS",
    "LeSpGEMM",
)


def _normalized(value: object) -> str:
    return "".join(ch for ch in str(value).strip().lower() if ch.isalnum())


def _find_column(headers: Sequence[object], name: str) -> int:
    target = _normalized(name)
    for column, header in enumerate(headers):
        if _normalized(header) == target:
            return column
    raise ValueError(f"Cannot find column {name!r}; worksheet headers: {list(headers)}")


def _to_float_or_nan(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number if math.isfinite(number) else math.nan


def _read_data(workbook_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Worksheet {sheet_name!r} does not exist; available sheets: "
                f"{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError(f"Worksheet {sheet_name!r} is empty") from exc

        columns = {
            "DATA_name": _find_column(headers, "DATA_name"),
            "Dense part": _find_column(headers, "Dense part"),
        }
        columns.update({name: _find_column(headers, name) for name in PERF_COLUMNS})

        data = []
        seen_names = set()
        for excel_row, row in enumerate(rows, start=2):
            if not any(value is not None for value in row):
                continue
            raw_name = row[columns["DATA_name"]]
            if raw_name is None or not str(raw_name).strip():
                raise ValueError(f"Missing DATA_name in Excel row {excel_row}")
            name = str(raw_name).strip()
            if name in seen_names:
                raise ValueError(f"Duplicate DATA_name {name!r}")
            seen_names.add(name)

            dense_part = _to_float_or_nan(row[columns["Dense part"]])
            record: dict[str, object] = {
                "DATA_name": name,
                "Dense part": dense_part,
            }
            for column in PERF_COLUMNS:
                record[column] = _to_float_or_nan(row[columns[column]])
            data.append(record)
    finally:
        workbook.close()

    if not data:
        raise ValueError(f"Worksheet {sheet_name!r} contains no benchmark rows")
    data.sort(
        key=lambda record: (
            not math.isfinite(float(record["Dense part"])),
            float(record["Dense part"])
            if math.isfinite(float(record["Dense part"]))
            else math.inf,
        )
    )
    return data


def _ellipsize(value: object, max_length: int = 8) -> str:
    text = str(value)
    if len(text) <= max_length:
        return text
    return text[: max_length - 3] + "..."


def _configure_fonts() -> None:
    plt.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [
                "Times New Roman",
                "Nimbus Roman",
                "Times",
                "Liberation Serif",
                "DejaVu Serif",
            ],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def _plot(data: Sequence[dict[str, object]], output_path: Path) -> None:
    _configure_fonts()

    names = [str(record["DATA_name"]) for record in data]
    x = np.arange(len(names))
    bar_width = 0.12
    offsets = (
        np.arange(len(PERF_COLUMNS)) - (len(PERF_COLUMNS) - 1) / 2.0
    ) * bar_width

    fig, ax = plt.subplots(figsize=(18, 3))
    for index, column in enumerate(PERF_COLUMNS):
        values = np.asarray([record[column] for record in data], dtype=float)
        finite = np.isfinite(values)
        positions = x + offsets[index]
        ax.bar(
            positions[finite],
            values[finite],
            width=bar_width,
            color=COLORS[index],
            label=LABELS[index],
            alpha=1.0,
            linewidth=0,
            zorder=10,
        )

        for x_position in positions[~finite]:
            ax.text(
                float(x_position),
                -0.8,
                "X",
                fontsize=10,
                color="darkred",
                ha="center",
                va="bottom",
                zorder=5,
                clip_on=False,
            )

    ax.set_xlabel(
        "Representative Matrices (the proportion of compact clusters in ascending order)",
        fontdict={"size": 16, "weight": "bold", "family": "serif"},
    )
    ax.set_ylabel(
        "Performance (GFlops)",
        fontdict={"size": 16, "weight": "bold", "family": "serif"},
    )
    ax.set_ylim(0, 30)

    le_index = PERF_COLUMNS.index("Le_GFlops")
    le_positions = x + offsets[le_index]
    target_names = {"gupta3", "TSOPF_FS_b300_c2"}
    y_text = ax.get_ylim()[1] - 0.3
    for row, name in enumerate(names):
        if name not in target_names:
            continue
        value = float(data[row]["Le_GFlops"])
        if math.isfinite(value):
            ax.text(
                float(le_positions[row]),
                min(value, y_text),
                f"{value:.2f}",
                fontsize=10,
                color="black",
                ha="center",
                va="bottom",
                zorder=11,
                clip_on=False,
            )

    x_padding = 2.0 * bar_width
    x_left = x[0] + offsets.min() - bar_width / 2.0 - x_padding
    x_right = x[-1] + offsets.max() + bar_width / 2.0 + x_padding
    ax.set_xlim(x_left, x_right)

    for boundary in range(len(x) - 1):
        ax.axvline(
            boundary + 0.5,
            color="0.55",
            linestyle=":",
            linewidth=0.9,
            alpha=0.75,
            zorder=1,
        )

    ax.set_xticks(x)
    ax.set_xticklabels([_ellipsize(name) for name in names], fontsize=12)
    plt.setp(ax.get_xticklabels(), fontweight="normal")
    plt.setp(ax.get_yticklabels(), fontsize=12, fontweight="bold")

    ax.legend(
        ncol=4,
        fontsize=12,
        frameon=True,
        loc="upper left",
        borderaxespad=0.12,
    )
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=5)
    fig.tight_layout()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, format="pdf", bbox_inches="tight")
    plt.close(fig)


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="Plot grouped SpGEMM performance for 20 representative matrices."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=script_dir / "20_perf.xlsx",
        help="input Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet",
        default="Xeon5120",
        help="worksheet containing performance data (default: %(default)s)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=script_dir / "20_perf.pdf",
        help="output PDF path (default: %(default)s)",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    data = _read_data(args.input, args.sheet)
    _plot(data, args.output)

    missing = sum(
        not math.isfinite(float(record[column]))
        for record in data
        for column in PERF_COLUMNS
    )
    print(f"Plotted {len(data)} matrices with {missing} missing bars")
    print(f"Saved {args.output.resolve()}")


if __name__ == "__main__":
    main()
